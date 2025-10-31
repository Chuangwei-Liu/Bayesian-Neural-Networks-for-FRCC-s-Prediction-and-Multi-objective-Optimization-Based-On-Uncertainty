import os
import numpy as np
import pandas as pd
import torch
from torch.utils.data import Dataset, DataLoader
from torch import nn
import torch.nn.functional as F
import torch.optim as optim
from sklearn.model_selection import KFold, train_test_split
from sklearn.metrics import mean_absolute_error, r2_score, mean_squared_error
import optuna
from optuna.samplers import TPESampler
from functools import partial
import matplotlib.pyplot as plt
import optuna.pruners as pruners
from optuna.pruners import MedianPruner
import functools
from torch.distributions import Normal
import math
from utils.Drawers import train_epoch_plot, scatter_plot, plot_prediction_vs_true
from utils.Data_worker import load_data, load_file, standard_scale_data, inverse_standard_scale_data
from utils.Setup_seed import setup_seed
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

OUTPUT_DIR = 'index_analysis_output'
EXCEL_PATH = OUTPUT_DIR + '/' + 'excel'
SCALER_PATH = OUTPUT_DIR + '/' + 'scaler'
FEATURE_SCALER_PATH = SCALER_PATH + '/' + 'feature_scaler'
TARGET_SCALER_PATH = SCALER_PATH + '/' + 'target_scaler'
MODEL_PATH = 'best_models'

DEVICE = torch.device('cpu')

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(FEATURE_SCALER_PATH, exist_ok=True)
os.makedirs(TARGET_SCALER_PATH, exist_ok=True)
os.makedirs(MODEL_PATH, exist_ok=True)
os.makedirs(EXCEL_PATH, exist_ok=True)

x_cols = ['SF/C', 'FA/C', 'W/B', 'S/B', 'SP/B', 'Vf(%)', 'If']
x_labels = ['SF/C', 'FA/C', 'W/B', 'S/B', 'SP/B', '$V_f$', '$I_f$']
y_cols = ['CX(MPa)', 'G(KJm3)', 'UTX(Mpa)', 'MiniSF(cm)']
y_labels = ['$\sigma_{cs}$', '$G_t$', '$\sigma_{u}$','$D_{spread}$']

x_label_mapping = dict(zip(x_cols, x_labels))
y_label_mapping = dict(zip(y_cols, y_labels))

raw_data = load_data(r'utils/Database.xlsx', sheet_names=['Sheet1'])

standard_scale_data(raw_data['Sheet1'][x_cols].values, train=True, 
                   scaler_file=f"{FEATURE_SCALER_PATH}/input_feature_scaler.pkl")
for y_col in y_cols:
    standard_scale_data(raw_data['Sheet1'][y_col].dropna().values.reshape(-1,1), 
                       train=True, scaler_file=f"{TARGET_SCALER_PATH}/{y_col}_target_scaler.pkl")
figs = ['Fig1', 'Fig2', 'Fig3']

class BayesianRegressor(nn.Module):
    def __init__(self, input_size, hidden_size=32, output_size=1, dropout_p=0.1):
        super().__init__()
        self.fc1 = nn.Linear(input_size, hidden_size)
        self.fc2 = nn.Linear(hidden_size, 2*hidden_size)
        self.output = nn.Linear(2*hidden_size, output_size*2)  # 输出均值和log方差
        self.dropout = nn.Dropout(p=dropout_p)

    def forward(self, x):
        x = F.relu(self.fc1(x))
        x = self.dropout(x)
        x = F.relu(self.fc2(x))
        x = self.dropout(x)
        output = self.output(x)
        
        # 将输出分成均值和log方差
        mean = output[:, :1]
        logvar = output[:, 1:]
        
        return mean, logvar

class BayesianNeuralNetwork(nn.Module):
    def __init__(self, input_size, hidden_size=32, output_size=1, dropout_p=0.1):

        super().__init__()
        self.model = BayesianRegressor(input_size, hidden_size, output_size, dropout_p)
        
    def forward(self, x):
        mean, logvar = self.model(x)
        return mean, logvar
    
    def predict_with_uncertainty(model, x, n_samples=100):
        # 数据不确定性 (eval模式)
        model.eval()
        mean, logvar = model(x)
        data_var = torch.exp(logvar).squeeze() 
        
        # 模型不确定性 (train模式)
        model.train()
        mc_samples = torch.stack([model(x)[0] for _ in range(n_samples)])
        model_var = mc_samples.var(dim=0).squeeze()  # 计算模型不确定性
        
        total_var = data_var + model_var
        return mean, total_var.sqrt(), data_var.sqrt(), model_var.sqrt()
    
    def nll(self, x, y):
        """计算负对数似然损失"""
        mean, logvar = self(x)
        inv_var = torch.exp(-logvar)
        loss = torch.sum(inv_var * (y - mean)**2 + logvar) / y.shape[0]
        return loss

# 数据集类（保持原样）
class Data(Dataset):
    def __init__(self, X, y):
        self.X = X
        self.y = y
        self.device = DEVICE

    def __len__(self):
        assert len(self.X) == len(self.y)
        return len(self.y)

    def __getitem__(self, i):
        return (
            torch.FloatTensor(self.X[i]).to(self.device),
            torch.FloatTensor(self.y[i]).to(self.device)
        )

if __name__ == "__main__":
    raw_data = load_data(r'utils/Index Analysis.xlsx', sheet_names=figs)
    feature_scaler = load_file(f"{FEATURE_SCALER_PATH}/input_feature_scaler.pkl")
    pred_columns = {
        'UTX(Mpa)': ['Predicted_UTX_mean', 'Predicted_UTX_total_std'],
        'G(KJm3)': ['Predicted_G_t_mean', 'Predicted_G_t_total_std'],
        'MiniSF(cm)': ['Predicted_D_spread_mean', 'Predicted_D_spread_total_std'],
        'CX(MPa)': ['Predicted_CX_mean', 'Predicted_CX_total_std']
    }
    
    best_model_list = {
        'UTX(Mpa)': 'UTX(Mpa)_424_best_model.pth',
        'G(KJm3)': "G(KJm3)_857_best_model.pth",
        'MiniSF(cm)': "MiniSF(cm)_114_best_model.pth",
        'CX(MPa)': "CX(MPa)_618_best_model.pth"  
    }

    excel_path = r'utils/Index Analysis.xlsx'
    wb = openpyxl.load_workbook(excel_path)

    for fig in figs:
        x_data = raw_data[fig][x_cols]
        x_scaled = feature_scaler.transform(x_data.values)
        
        # 获取当前sheet
        sheet = wb[fig]
        
        # 查找"Ef"列的位置
        ef_col = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == "Ef":
                ef_col = col
                break
        
        # 如果没找到Ef列，则从最后一列开始添加
        start_col = ef_col + 1 if ef_col else sheet.max_column + 1
        next_col = start_col
        
        for y_col in y_cols:
            print(f"Processing {y_col}...")
            target_scaler = load_file(f"{TARGET_SCALER_PATH}/{y_col}_target_scaler.pkl")
            model_path = os.path.join(MODEL_PATH, best_model_list[y_col])
            if not os.path.exists(model_path):
                print(f"⚠️ Model file not found: {model_path}")
                continue
                
            model = torch.load(model_path, map_location=DEVICE, weights_only=False)
            mean, total_std, data_std, model_std = BayesianNeuralNetwork.predict_with_uncertainty(
                model, torch.FloatTensor(x_scaled).to(DEVICE), n_samples=100
                )

            mean_np = mean.detach().cpu().numpy().reshape(-1, 1)
            total_std_np = total_std.detach().cpu().numpy()

            mean_np = target_scaler.inverse_transform(mean_np).flatten()
            total_std_np = total_std_np * (target_scaler.data_max_ - target_scaler.data_min_)
                
            pred_mean_col, pred_std_col = pred_columns[y_col]
                
            sheet.cell(row=1, column=next_col, value=pred_mean_col)
            sheet.cell(row=1, column=next_col+1, value=pred_std_col)
                
            # 写入数据
            for i in range(len(mean_np)):
                sheet.cell(row=i+2, column=next_col, value=float(mean_np[i]))
                sheet.cell(row=i+2, column=next_col+1, value=float(total_std_np[i]))
                
            next_col += 2  # 为下一组预测留出两列
            print(f"✓ Successfully processed {y_col}")

    output_path = EXCEL_PATH + "/" + "index_analysis_pred.xlsx"
    wb.save(output_path)
    print(f"预测结果已保存到: {output_path}")